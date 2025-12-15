"""
Módulo de Exportación para Forecast MR

Proporciona funciones para exportar resultados de forecast
a diferentes formatos: PDF, CSV, Excel.

Author: Manuel Remón
"""

from typing import Dict, Any, Optional, List
import pandas as pd
import numpy as np
from datetime import datetime
from io import BytesIO, StringIO
import base64
from loguru import logger

# Importaciones condicionales para PDF
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch, cm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        Image, PageBreak, HRFlowable
    )
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    logger.warning("reportlab no disponible para exportación PDF")

# Importaciones para Excel
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl no disponible para exportación Excel avanzada")


class ForecastExporter:
    """
    Exportador de resultados de forecast.

    Soporta múltiples formatos:
    - PDF: Reporte profesional con gráficos y tablas
    - CSV: Datos tabulares simples
    - Excel: Múltiples hojas con formato

    Ejemplo de uso:
        exporter = ForecastExporter()

        # Exportar a PDF
        pdf_bytes = exporter.exportar_pdf(
            predicciones=df_pred,
            metricas=metrics,
            material_info={'codigo': '123', 'descripcion': 'Material X'}
        )

        # Exportar a CSV
        csv_string = exporter.exportar_csv(df_pred)

        # Exportar a Excel
        excel_bytes = exporter.exportar_excel(df_pred, metricas)
    """

    def __init__(self):
        """Inicializa el exportador"""
        self.styles = None

        # Colores corporativos (solo si reportlab disponible)
        if REPORTLAB_AVAILABLE:
            self.COLOR_PRIMARIO = colors.HexColor('#1e293b')
            self.COLOR_SECUNDARIO = colors.HexColor('#3b82f6')
            self.COLOR_EXITO = colors.HexColor('#22c55e')
            self.COLOR_ALERTA = colors.HexColor('#f59e0b')
            self.COLOR_ERROR = colors.HexColor('#ef4444')
            self._setup_styles()
        else:
            # Colores placeholder para cuando reportlab no está disponible
            self.COLOR_PRIMARIO = None
            self.COLOR_SECUNDARIO = None
            self.COLOR_EXITO = None
            self.COLOR_ALERTA = None
            self.COLOR_ERROR = None

    def _setup_styles(self):
        """Configura estilos para PDF"""
        self.styles = getSampleStyleSheet()

        # Estilo título principal
        self.styles.add(ParagraphStyle(
            name='TituloPrincipal',
            parent=self.styles['Heading1'],
            fontSize=18,
            spaceAfter=20,
            textColor=self.COLOR_PRIMARIO,
            alignment=TA_CENTER
        ))

        # Estilo subtítulo
        self.styles.add(ParagraphStyle(
            name='Subtitulo',
            parent=self.styles['Heading2'],
            fontSize=14,
            spaceAfter=12,
            textColor=self.COLOR_SECUNDARIO
        ))

        # Estilo texto normal
        self.styles.add(ParagraphStyle(
            name='TextoNormal',
            parent=self.styles['Normal'],
            fontSize=10,
            spaceAfter=6
        ))

        # Estilo métrica
        self.styles.add(ParagraphStyle(
            name='Metrica',
            parent=self.styles['Normal'],
            fontSize=12,
            alignment=TA_CENTER,
            textColor=self.COLOR_PRIMARIO
        ))

    def exportar_pdf(
        self,
        predicciones: pd.DataFrame,
        metricas: Dict[str, float],
        material_info: Dict[str, str],
        config: Optional[Dict[str, Any]] = None,
        validacion: Optional[Dict[str, Any]] = None,
        incluir_grafico: bool = False,
        grafico_bytes: Optional[bytes] = None
    ) -> bytes:
        """
        Exporta resultados a PDF profesional.

        Args:
            predicciones: DataFrame con predicciones
            metricas: Diccionario con métricas del modelo
            material_info: Info del material (codigo, descripcion)
            config: Configuración usada
            validacion: Resultado de validación de datos
            incluir_grafico: Si incluir gráfico
            grafico_bytes: Bytes de imagen del gráfico

        Returns:
            Bytes del PDF generado
        """
        if not REPORTLAB_AVAILABLE:
            raise ImportError("reportlab es requerido para exportar PDF")

        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )

        elementos = []

        # Header
        elementos.extend(self._crear_header_pdf(material_info))

        # Resumen de métricas
        elementos.extend(self._crear_metricas_pdf(metricas))

        # Configuración usada
        if config:
            elementos.extend(self._crear_config_pdf(config))

        # Validación de datos
        if validacion:
            elementos.extend(self._crear_validacion_pdf(validacion))

        # Gráfico
        if incluir_grafico and grafico_bytes:
            elementos.extend(self._crear_grafico_pdf(grafico_bytes))

        # Tabla de predicciones
        elementos.extend(self._crear_tabla_predicciones_pdf(predicciones))

        # Footer
        elementos.extend(self._crear_footer_pdf())

        # Generar PDF
        doc.build(elementos)
        buffer.seek(0)

        logger.info("PDF exportado exitosamente")
        return buffer.getvalue()

    def _crear_header_pdf(self, material_info: Dict[str, str]) -> List:
        """Crea header del PDF"""
        elementos = []

        # Título
        elementos.append(Paragraph(
            "Forecast MR - Reporte de Predicción",
            self.styles['TituloPrincipal']
        ))

        # Información del material
        codigo = material_info.get('codigo', 'N/A')
        descripcion = material_info.get('descripcion', 'Sin descripción')

        elementos.append(Paragraph(
            f"<b>Material:</b> {codigo}",
            self.styles['TextoNormal']
        ))
        elementos.append(Paragraph(
            f"<b>Descripción:</b> {descripcion}",
            self.styles['TextoNormal']
        ))
        elementos.append(Paragraph(
            f"<b>Fecha de generación:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}",
            self.styles['TextoNormal']
        ))

        elementos.append(Spacer(1, 20))
        elementos.append(HRFlowable(width="100%", thickness=1, color=self.COLOR_SECUNDARIO))
        elementos.append(Spacer(1, 20))

        return elementos

    def _crear_metricas_pdf(self, metricas: Dict[str, float]) -> List:
        """Crea sección de métricas"""
        elementos = []

        elementos.append(Paragraph("Métricas del Modelo", self.styles['Subtitulo']))

        # Tabla de métricas
        data = [
            ['Métrica', 'Valor', 'Interpretación'],
            ['R² (Precisión)', f"{metricas.get('r2', 0)*100:.1f}%", self._interpretar_r2(metricas.get('r2', 0))],
            ['MAE (Error Medio)', f"{metricas.get('mae', 0):.2f}", 'Error absoluto promedio'],
            ['RMSE', f"{metricas.get('rmse', 0):.2f}", 'Penaliza errores grandes'],
            ['MAPE', f"{metricas.get('mape', 0):.1f}%", 'Error porcentual promedio']
        ]

        tabla = Table(data, colWidths=[4*cm, 3*cm, 8*cm])
        tabla.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), self.COLOR_PRIMARIO),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fafc')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ]))

        elementos.append(tabla)
        elementos.append(Spacer(1, 20))

        return elementos

    def _interpretar_r2(self, r2: float) -> str:
        """Interpreta valor de R²"""
        if r2 >= 0.9:
            return "Excelente precisión"
        elif r2 >= 0.7:
            return "Buena precisión"
        elif r2 >= 0.5:
            return "Precisión moderada"
        else:
            return "Precisión mejorable"

    def _crear_config_pdf(self, config: Dict[str, Any]) -> List:
        """Crea sección de configuración"""
        elementos = []

        elementos.append(Paragraph("Configuración", self.styles['Subtitulo']))

        data = [['Parámetro', 'Valor']]
        for key, value in config.items():
            if key not in ['material']:  # Excluir algunos
                data.append([key.replace('_', ' ').title(), str(value)])

        if len(data) > 1:
            tabla = Table(data, colWidths=[6*cm, 9*cm])
            tabla.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), self.COLOR_SECUNDARIO),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            elementos.append(tabla)

        elementos.append(Spacer(1, 20))
        return elementos

    def _crear_validacion_pdf(self, validacion: Dict[str, Any]) -> List:
        """Crea sección de validación"""
        elementos = []

        elementos.append(Paragraph("Calidad de Datos", self.styles['Subtitulo']))

        score = validacion.get('score', 0)
        color_score = self.COLOR_EXITO if score >= 70 else (self.COLOR_ALERTA if score >= 50 else self.COLOR_ERROR)

        elementos.append(Paragraph(
            f"<b>Score de Calidad:</b> {score:.0f}/100",
            self.styles['TextoNormal']
        ))

        if validacion.get('resumen'):
            elementos.append(Paragraph(
                validacion['resumen'],
                self.styles['TextoNormal']
            ))

        elementos.append(Spacer(1, 20))
        return elementos

    def _crear_grafico_pdf(self, grafico_bytes: bytes) -> List:
        """Crea sección de gráfico"""
        elementos = []

        elementos.append(Paragraph("Gráfico de Predicción", self.styles['Subtitulo']))

        # Convertir bytes a imagen
        img_buffer = BytesIO(grafico_bytes)
        img = Image(img_buffer, width=15*cm, height=8*cm)
        elementos.append(img)

        elementos.append(Spacer(1, 20))
        return elementos

    def _crear_tabla_predicciones_pdf(self, predicciones: pd.DataFrame) -> List:
        """Crea tabla de predicciones"""
        elementos = []

        elementos.append(Paragraph("Detalle de Predicciones", self.styles['Subtitulo']))

        # Preparar datos (limitar a primeras 30 filas)
        df_mostrar = predicciones.head(30).copy()

        # Formatear columnas
        columnas = ['fecha', 'prediccion']
        if 'limite_inferior' in df_mostrar.columns:
            columnas.extend(['limite_inferior', 'limite_superior'])
        if 'dia_semana' in df_mostrar.columns:
            columnas.append('dia_semana')

        df_mostrar = df_mostrar[[c for c in columnas if c in df_mostrar.columns]]

        # Formatear valores
        if 'fecha' in df_mostrar.columns:
            df_mostrar['fecha'] = pd.to_datetime(df_mostrar['fecha']).dt.strftime('%d/%m/%Y')
        for col in ['prediccion', 'limite_inferior', 'limite_superior']:
            if col in df_mostrar.columns:
                df_mostrar[col] = df_mostrar[col].apply(lambda x: f"{x:,.1f}" if pd.notna(x) else "-")

        # Crear tabla
        data = [df_mostrar.columns.tolist()] + df_mostrar.values.tolist()

        tabla = Table(data, repeatRows=1)
        tabla.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), self.COLOR_PRIMARIO),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')])
        ]))

        elementos.append(tabla)

        if len(predicciones) > 30:
            elementos.append(Spacer(1, 10))
            elementos.append(Paragraph(
                f"<i>Mostrando primeras 30 de {len(predicciones)} predicciones</i>",
                self.styles['TextoNormal']
            ))

        elementos.append(Spacer(1, 20))
        return elementos

    def _crear_footer_pdf(self) -> List:
        """Crea footer del PDF"""
        elementos = []

        elementos.append(HRFlowable(width="100%", thickness=1, color=self.COLOR_SECUNDARIO))
        elementos.append(Spacer(1, 10))

        elementos.append(Paragraph(
            "Generado por Forecast MR v1.0 | Machine Learning para Predicción de Demanda",
            ParagraphStyle(
                'Footer',
                parent=self.styles['Normal'],
                fontSize=8,
                textColor=colors.gray,
                alignment=TA_CENTER
            )
        ))
        elementos.append(Paragraph(
            "Manuel Remón | Neuquén, Argentina",
            ParagraphStyle(
                'FooterAuthor',
                parent=self.styles['Normal'],
                fontSize=8,
                textColor=colors.gray,
                alignment=TA_CENTER
            )
        ))

        return elementos

    def exportar_csv(
        self,
        predicciones: pd.DataFrame,
        incluir_metricas: bool = False,
        metricas: Optional[Dict[str, float]] = None
    ) -> str:
        """
        Exporta predicciones a CSV.

        Args:
            predicciones: DataFrame con predicciones
            incluir_metricas: Si incluir métricas como comentario
            metricas: Diccionario con métricas

        Returns:
            String con contenido CSV
        """
        output = StringIO()

        # Header con métricas como comentario
        if incluir_metricas and metricas:
            output.write(f"# Forecast MR - Exportación {datetime.now().strftime('%Y-%m-%d %H:%M')}\n")
            output.write(f"# R2: {metricas.get('r2', 0):.4f}\n")
            output.write(f"# MAE: {metricas.get('mae', 0):.2f}\n")
            output.write(f"# RMSE: {metricas.get('rmse', 0):.2f}\n")
            output.write("#\n")

        # Datos
        predicciones.to_csv(output, index=False)

        logger.info("CSV exportado exitosamente")
        return output.getvalue()

    def exportar_excel(
        self,
        predicciones: pd.DataFrame,
        metricas: Optional[Dict[str, float]] = None,
        material_info: Optional[Dict[str, str]] = None,
        historico: Optional[pd.DataFrame] = None
    ) -> bytes:
        """
        Exporta a Excel con múltiples hojas.

        Args:
            predicciones: DataFrame con predicciones
            metricas: Diccionario con métricas
            material_info: Info del material
            historico: Datos históricos opcionales

        Returns:
            Bytes del archivo Excel
        """
        buffer = BytesIO()

        with pd.ExcelWriter(buffer, engine='openpyxl' if OPENPYXL_AVAILABLE else 'xlsxwriter') as writer:
            # Hoja de predicciones
            predicciones.to_excel(writer, sheet_name='Predicciones', index=False)

            # Hoja de métricas
            if metricas:
                df_metricas = pd.DataFrame([
                    {'Métrica': 'R² (Precisión)', 'Valor': f"{metricas.get('r2', 0)*100:.1f}%"},
                    {'Métrica': 'MAE (Error Medio)', 'Valor': f"{metricas.get('mae', 0):.2f}"},
                    {'Métrica': 'RMSE', 'Valor': f"{metricas.get('rmse', 0):.2f}"},
                    {'Métrica': 'MAPE', 'Valor': f"{metricas.get('mape', 0):.1f}%"}
                ])
                df_metricas.to_excel(writer, sheet_name='Métricas', index=False)

            # Hoja de info
            if material_info:
                df_info = pd.DataFrame([
                    {'Campo': 'Código', 'Valor': material_info.get('codigo', '')},
                    {'Campo': 'Descripción', 'Valor': material_info.get('descripcion', '')},
                    {'Campo': 'Fecha Generación', 'Valor': datetime.now().strftime('%Y-%m-%d %H:%M')},
                    {'Campo': 'Aplicación', 'Valor': 'Forecast MR v1.0'}
                ])
                df_info.to_excel(writer, sheet_name='Información', index=False)

            # Hoja de histórico
            if historico is not None:
                historico.to_excel(writer, sheet_name='Histórico', index=False)

            # Aplicar formato si openpyxl disponible
            if OPENPYXL_AVAILABLE:
                self._aplicar_formato_excel(writer)

        buffer.seek(0)
        logger.info("Excel exportado exitosamente")
        return buffer.getvalue()

    def _aplicar_formato_excel(self, writer):
        """Aplica formato a hojas Excel"""
        workbook = writer.book

        # Estilo header
        header_fill = PatternFill(start_color='1e293b', end_color='1e293b', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]

            # Formato header
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

            # Ajustar ancho columnas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)


# Funciones de conveniencia
def exportar_forecast_pdf(
    predicciones: pd.DataFrame,
    metricas: Dict[str, float],
    material: str,
    descripcion: str = ""
) -> bytes:
    """
    Función de conveniencia para exportar PDF.

    Args:
        predicciones: DataFrame con predicciones
        metricas: Métricas del modelo
        material: Código del material
        descripcion: Descripción del material

    Returns:
        Bytes del PDF
    """
    exporter = ForecastExporter()
    return exporter.exportar_pdf(
        predicciones=predicciones,
        metricas=metricas,
        material_info={'codigo': material, 'descripcion': descripcion}
    )


def exportar_forecast_csv(predicciones: pd.DataFrame) -> str:
    """
    Función de conveniencia para exportar CSV.

    Args:
        predicciones: DataFrame con predicciones

    Returns:
        String CSV
    """
    exporter = ForecastExporter()
    return exporter.exportar_csv(predicciones)


def exportar_forecast_excel(
    predicciones: pd.DataFrame,
    metricas: Optional[Dict[str, float]] = None
) -> bytes:
    """
    Función de conveniencia para exportar Excel.

    Args:
        predicciones: DataFrame con predicciones
        metricas: Métricas opcionales

    Returns:
        Bytes del Excel
    """
    exporter = ForecastExporter()
    return exporter.exportar_excel(predicciones, metricas)


def crear_download_link(data: bytes, filename: str, mime_type: str) -> str:
    """
    Crea link de descarga para Dash.

    Args:
        data: Bytes del archivo
        filename: Nombre del archivo
        mime_type: Tipo MIME

    Returns:
        String href para dcc.Download
    """
    b64 = base64.b64encode(data).decode()
    return f"data:{mime_type};base64,{b64}"
